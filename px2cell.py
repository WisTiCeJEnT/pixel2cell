import cv2
import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell   

def findHex(c):
    ls = []
    for i in range (len(c)):
        for j in range (len(c[i])):
            color_code = (str(hex(int(c[i][j])))[2:])
            if len(color_code) != 2:
                color_code = '0' + color_code
            ls.append(color_code)
    return ls

def mergeHex(r,g,b):
    res_ls = []
    for i in range (len(findHex(r))):
        res_ls.append('FF' + red[i] + green[i] + blue[i])
    return res_ls

def colnum_string(m,n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return (string + str(m))


def colorFill(cell,color):
    cellFill = PatternFill(start_color=color,
                   end_color=color,
                   fill_type='solid')
    cell.fill = cellFill

wb = Workbook()
ws = wb.active

img = cv2.imread('./dota2.png')
height, width = img.shape[:2]
b,g,r = cv2.split(img)
cv2.imshow('r',r)
cv2.imshow('g',g)
cv2.imshow('b',b)

red = findHex(r)
green = findHex(g)
blue = findHex(b)

resMerge = np.reshape(mergeHex(r,g,b), (-1, width))

row_count = 0
col_count = 0

for col in ws.iter_cols(min_col=1, max_col=width, min_row=1, max_row=height):
    for cell in col:
        ws[(colnum_string(row_count+1,col_count+1))] = ' '
        colorFill(cell,resMerge[row_count][col_count].upper())
        row_count += 1
    col_count += 1
    row_count = 0
wb.save('test.xlsx')
